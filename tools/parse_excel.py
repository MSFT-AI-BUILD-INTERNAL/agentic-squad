#!/usr/bin/env python3
"""Excel Parser Tool for Copilot CLI.

Parses .xlsx files and outputs structured data (JSON or Markdown table)
that Copilot CLI agents can consume and reason about.

Usage:
    python tools/parse_excel.py <file.xlsx> [options]

Options:
    --sheet NAME       Sheet name to parse (default: active sheet)
    --sheets           List all sheet names
    --format FORMAT    Output format: json, markdown, csv (default: json)
    --range RANGE      Cell range to parse, e.g. A1:D10
    --header ROW       Row number to use as header (default: 1, 0 for no header)
    --max-rows N       Maximum rows to output (default: unlimited)
    --summary          Output summary statistics instead of full data

Examples:
    python tools/parse_excel.py data.xlsx
    python tools/parse_excel.py data.xlsx --sheet "Sheet2" --format markdown
    python tools/parse_excel.py data.xlsx --range A1:C20 --format csv
    python tools/parse_excel.py data.xlsx --summary
    python tools/parse_excel.py data.xlsx --sheets
"""

import argparse
import json
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.utils import range_boundaries
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def parse_args():
    parser = argparse.ArgumentParser(
        description="Parse Excel (.xlsx) files for Copilot CLI consumption"
    )
    parser.add_argument("file", help="Path to the .xlsx file")
    parser.add_argument("--sheet", help="Sheet name to parse (default: active sheet)")
    parser.add_argument("--sheets", action="store_true", help="List all sheet names")
    parser.add_argument(
        "--format",
        choices=["json", "markdown", "csv"],
        default="json",
        help="Output format (default: json)",
    )
    parser.add_argument("--range", help="Cell range to parse, e.g. A1:D10")
    parser.add_argument(
        "--header",
        type=int,
        default=1,
        help="Row number for header (default: 1, 0 for no header)",
    )
    parser.add_argument("--max-rows", type=int, help="Maximum rows to output")
    parser.add_argument(
        "--summary", action="store_true", help="Output summary statistics"
    )
    return parser.parse_args()


def load_sheet(wb, sheet_name=None):
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            print(
                f"Error: Sheet '{sheet_name}' not found. "
                f"Available: {', '.join(wb.sheetnames)}",
                file=sys.stderr,
            )
            sys.exit(1)
        return wb[sheet_name]
    return wb.active


def cell_value(cell):
    """Convert cell value to a JSON-serializable type."""
    if cell.value is None:
        return None
    if hasattr(cell.value, "isoformat"):
        return cell.value.isoformat()
    return cell.value


def extract_data(ws, cell_range=None, header_row=1, max_rows=None):
    """Extract data from worksheet as list of dicts (or list of lists if no header)."""
    if cell_range:
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    else:
        min_row = ws.min_row
        max_row_ws = ws.max_row
        min_col = ws.min_column
        max_col = ws.max_column
        max_row = max_row_ws

    if max_rows and header_row:
        max_row = min(max_row, (header_row) + max_rows)
    elif max_rows:
        max_row = min(max_row, min_row + max_rows - 1)

    rows = []
    for row in ws.iter_rows(
        min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
    ):
        rows.append([cell_value(c) for c in row])

    if not rows:
        return [], []

    if header_row and header_row >= min_row:
        header_idx = header_row - min_row
        if header_idx < len(rows):
            headers = [str(h) if h else f"col_{i}" for i, h in enumerate(rows[header_idx])]
            data_rows = rows[header_idx + 1 :]
            records = []
            for row in data_rows:
                record = {}
                for i, val in enumerate(row):
                    key = headers[i] if i < len(headers) else f"col_{i}"
                    record[key] = val
                records.append(record)
            return headers, records
    
    headers = [f"col_{i}" for i in range(len(rows[0]))]
    records = [{headers[i]: val for i, val in enumerate(row)} for row in rows]
    return headers, records


def format_json(headers, records, sheet_name):
    output = {
        "sheet": sheet_name,
        "total_rows": len(records),
        "columns": headers,
        "data": records,
    }
    return json.dumps(output, ensure_ascii=False, indent=2, default=str)


def format_markdown(headers, records, sheet_name):
    if not headers:
        return f"*Sheet '{sheet_name}' is empty.*"

    lines = [f"## Sheet: {sheet_name}", ""]
    lines.append("| " + " | ".join(headers) + " |")
    lines.append("| " + " | ".join(["---"] * len(headers)) + " |")

    for record in records:
        values = [str(record.get(h, "")) if record.get(h) is not None else "" for h in headers]
        lines.append("| " + " | ".join(values) + " |")

    lines.append("")
    lines.append(f"*{len(records)} rows*")
    return "\n".join(lines)


def format_csv(headers, records):
    import csv
    import io

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=headers)
    writer.writeheader()
    writer.writerows(records)
    return output.getvalue()


def generate_summary(ws, headers, records, sheet_name):
    """Generate summary statistics about the data."""
    summary = {
        "sheet": sheet_name,
        "total_rows": len(records),
        "total_columns": len(headers),
        "columns": {},
    }

    for h in headers:
        values = [r.get(h) for r in records if r.get(h) is not None]
        col_info = {
            "non_null_count": len(values),
            "null_count": len(records) - len(values),
        }

        numeric_values = [v for v in values if isinstance(v, (int, float))]
        if numeric_values:
            col_info["type"] = "numeric"
            col_info["min"] = min(numeric_values)
            col_info["max"] = max(numeric_values)
            col_info["avg"] = round(sum(numeric_values) / len(numeric_values), 2)
        else:
            col_info["type"] = "text"
            str_values = [str(v) for v in values]
            if str_values:
                unique = set(str_values)
                col_info["unique_count"] = len(unique)
                if len(unique) <= 10:
                    col_info["unique_values"] = sorted(unique)
                col_info["sample"] = str_values[:3]

        summary["columns"][h] = col_info

    return json.dumps(summary, ensure_ascii=False, indent=2, default=str)


def main():
    args = parse_args()

    filepath = Path(args.file)
    if not filepath.exists():
        print(f"Error: File not found: {filepath}", file=sys.stderr)
        sys.exit(1)

    if not filepath.suffix.lower() in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        print(
            f"Error: Unsupported file format '{filepath.suffix}'. "
            "Supported: .xlsx, .xlsm, .xltx, .xltm",
            file=sys.stderr,
        )
        sys.exit(1)

    wb = load_workbook(filepath, read_only=True, data_only=True)

    if args.sheets:
        result = {"sheets": wb.sheetnames, "active_sheet": wb.active.title}
        print(json.dumps(result, ensure_ascii=False, indent=2))
        wb.close()
        return

    ws = load_sheet(wb, args.sheet)
    sheet_name = ws.title

    headers, records = extract_data(
        ws,
        cell_range=args.range,
        header_row=args.header,
        max_rows=args.max_rows,
    )

    if args.summary:
        print(generate_summary(ws, headers, records, sheet_name))
    elif args.format == "json":
        print(format_json(headers, records, sheet_name))
    elif args.format == "markdown":
        print(format_markdown(headers, records, sheet_name))
    elif args.format == "csv":
        print(format_csv(headers, records))

    wb.close()


if __name__ == "__main__":
    main()
